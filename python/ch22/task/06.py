import openpyxl
from openpyxl import Workbook,load_workbook
from datetime import datetime

# 파일 만들기
wb = Workbook()
sheet = wb.active
sheet['A1']='Hello'
sheet['B1']='World'
wb.save(r'.\python\ch22\task\new_example.xlsx')

# 파일 가져오기/내용 수정
wb = load_workbook(r'.\python\ch22\task\new_example.xlsx')
sheet = wb.active
print(sheet['A1'].value)
print(sheet['B1'].value)
sheet['C1']= 'Python'
sheet['A2']= datetime(2025,10,16)
sheet['B2']= '=B1*2'

wb.save(r'.\python\ch22\task\new_example.xlsx')

for row in sheet.iter_rows(min_row=1,max_col=3,max_row=5):
    for cell in row:
        print(cell.value)

