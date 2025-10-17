import openpyxl
from openpyxl import Workbook,load_workbook

# 1. 엑셀 파일을 생성하고 기본 데이터를 입력하는 함수
def create_excel_file(file_path,data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sample Data"

    for row in data:
        ws.append(row)
    wb.save(file_path)
    return ws

# 2. 기존 엑셀 파일을 열어 데이터를 읽고 출력하는 함수
def readExcel(sheet,max_row,max_col):
    for row in sheet.iter_rows(max_row=max_row,max_col=max_col):
        for cell in row:
            print(cell.value,end="\t")
        print()

#3. 함수 호출
if __name__=="__main__":
    data_list=[
        ["이름","나이","국가"],
        ["홍길동",20,"대한민국"],
        ["Alice",25,"USA"]
    ]
    excel_path=r'.\python\ch22\task\sample.xlsx'
    sheet = create_excel_file(excel_path,data_list)
    readExcel(sheet,3,3)