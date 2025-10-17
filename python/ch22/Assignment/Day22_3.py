import openpyxl
from openpyxl import Workbook,load_workbook

wb=Workbook()
sheet=wb.active
wb.save(r".\python\ch22\Assignment\data.xlsx")

wb=load_workbook()
sheet=wb.active
sheet['A1']='Hello'
wb.save(r".\python\ch22\Assignment\data.xlsx")